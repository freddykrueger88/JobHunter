"""Daten-Export und -Import (DSGVO Art. 20 – Datenportabilität).
#65: CSV- und XLSX-Export hinzugefügt.
"""
import json
import io
import csv
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from backend.core.database import get_db
from backend.models.application import Application
from backend.models.job import Job
from backend.models.cv import CVData
from backend.models.history import HistoryEntry
from backend.models.reminder import Reminder
from backend.models.settings import UserSettings

router = APIRouter(prefix="/export", tags=["Export/Import"])
EXPORT_VERSION = "1.2"


async def _serialize(obj) -> dict:
    d = {c.name: getattr(obj, c.name) for c in obj.__table__.columns}
    for k, v in d.items():
        if isinstance(v, datetime):
            d[k] = v.isoformat()
    return d


@router.get("/")
async def export_all(db: AsyncSession = Depends(get_db)):
    """Vollständiger JSON-Export aller Daten."""
    data = {
        "version": EXPORT_VERSION,
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "jobs": [await _serialize(r) for r in (await db.execute(select(Job))).scalars().all()],
        "applications": [await _serialize(r) for r in (await db.execute(select(Application))).scalars().all()],
        "reminders": [await _serialize(r) for r in (await db.execute(select(Reminder))).scalars().all()],
        "history": [await _serialize(r) for r in (await db.execute(select(HistoryEntry))).scalars().all()],
        "cvs": [await _serialize(r) for r in (await db.execute(select(CVData))).scalars().all()],
    }
    json_bytes = json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
    filename = f"jobhunter_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    return StreamingResponse(
        io.BytesIO(json_bytes),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/csv")
async def export_csv(db: AsyncSession = Depends(get_db)):
    """#65 – Bewerbungen + Jobs als CSV."""
    apps = (await db.execute(select(Application).order_by(Application.created_at.desc()))).scalars().all()
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_ALL)
    writer.writerow(["ID", "Firma", "Stelle", "Ort", "Status", "Beworben am", "Gespräch am", "Erstellt am", "Notizen"])
    for a in apps:
        job = await db.get(Job, a.job_id)
        writer.writerow([
            a.id,
            job.company if job else "",
            job.title if job else "",
            job.location if job else "",
            a.status,
            a.applied_at.strftime("%d.%m.%Y") if a.applied_at else "",
            a.interview_at.strftime("%d.%m.%Y %H:%M") if a.interview_at else "",
            a.created_at.strftime("%d.%m.%Y") if a.created_at else "",
            a.notes or "",
        ])
    filename = f"jobhunter_bewerbungen_{datetime.now().strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        io.BytesIO(("\ufeff" + buf.getvalue()).encode("utf-8")),  # BOM für Excel
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/xlsx")
async def export_xlsx(db: AsyncSession = Depends(get_db)):
    """#65 – Bewerbungen als Excel-Datei (openpyxl)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl nicht installiert (pip install openpyxl)")

    apps = (await db.execute(select(Application).order_by(Application.created_at.desc()))).scalars().all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bewerbungen"

    headers = ["ID", "Firma", "Stelle", "Ort", "Status", "Beworben am", "Gespräch am", "Erstellt am", "Notizen"]
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    STATUS_COLORS = {
        "interessant": "D6EAF8",
        "beworben": "D5F5E3",
        "interview": "FDEBD0",
        "angenommen": "A9DFBF",
        "absage": "FADBD8",
        "archiviert": "EAECEE",
    }

    for row_i, a in enumerate(apps, 2):
        job = await db.get(Job, a.job_id)
        row = [
            a.id,
            job.company if job else "",
            job.title if job else "",
            job.location if job else "",
            a.status,
            a.applied_at.strftime("%d.%m.%Y") if a.applied_at else "",
            a.interview_at.strftime("%d.%m.%Y %H:%M") if a.interview_at else "",
            a.created_at.strftime("%d.%m.%Y") if a.created_at else "",
            a.notes or "",
        ]
        color = STATUS_COLORS.get(a.status, "FFFFFF")
        for col_i, val in enumerate(row, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.fill = PatternFill("solid", fgColor=color)

    # Spaltenbreiten anpassen
    col_widths = [6, 25, 30, 18, 14, 14, 18, 14, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"jobhunter_bewerbungen_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/import")
async def import_data(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Importiert einen JobHunter-JSON-Export. Bestehende Daten werden nicht überschrieben."""
    if not file.filename.endswith(".json"):
        raise HTTPException(status_code=400, detail="Nur .json-Dateien erlaubt")
    raw = await file.read()
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Ungültiges JSON")

    version = data.get("version", "unknown")
    stats = {"jobs": 0, "applications": 0, "reminders": 0, "history": 0}

    for item in data.get("jobs", []):
        existing = None
        if item.get("external_id"):
            res = await db.execute(select(Job).where(Job.external_id == item["external_id"]))
            existing = res.scalar_one_or_none()
        if not existing:
            job = Job(**{k: v for k, v in item.items() if k != "id" and hasattr(Job, k)})
            db.add(job)
            stats["jobs"] += 1

    for item in data.get("reminders", []):
        r = Reminder(**{k: v for k, v in item.items() if k != "id" and hasattr(Reminder, k)})
        db.add(r)
        stats["reminders"] += 1

    for item in data.get("history", []):
        h = HistoryEntry(**{k: v for k, v in item.items() if k != "id" and hasattr(HistoryEntry, k)})
        db.add(h)
        stats["history"] += 1

    await db.commit()
    db.add(HistoryEntry(
        event_type="data_imported",
        description=f"Import aus Version {version}: {stats['jobs']} Stellen, {stats['reminders']} Erinnerungen",
        meta=stats,
    ))
    await db.commit()
    return {"imported": stats, "source_version": version}
